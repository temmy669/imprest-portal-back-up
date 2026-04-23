import logging
from collections import Counter
from rest_framework.generics import get_object_or_404
from utils.pagination import DynamicPageSizePagination
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import *
from .serializers import *
from purchases.models import PurchaseRequest, LimitConfig, PurchaseRequestItem
from utils.permissions import (ViewReimbursementRequest,
                               SubmitReimbursementRequest,
                               ApproveReimbursementRequest,
                               DeclineReimbursementRequest,
                               ChangeReimbursementRequest,
                                DisburseReimbursementRequest)
from helpers.response import CustomResponse
from users.auth import JWTAuthenticationFromCookie
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.files.storage import FileSystemStorage
import os
from django.conf import settings
from django.utils import timezone
from decimal import InvalidOperation, Decimal

from django.core.files.storage import FileSystemStorage
from rest_framework.parsers import MultiPartParser, FormParser

from django.db.models import Q, Count, Sum
from collections import Counter
from drf_spectacular.utils import extend_schema, OpenApiParameter
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

import cloudinary
import cloudinary.uploader
import re
from django.db import transaction
from utils.email_utils import send_reimbursement_rejection_notification, send_reimbursement_approval_notification
from .post_to_byd import update_sap_record
from roles.models import Role

logger = logging.getLogger(__name__)

class ReimbursementRequestView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated(), ViewReimbursementRequest()]
        elif self.request.method == 'POST':
            return [IsAuthenticated(), SubmitReimbursementRequest()]
        elif self.request.method == 'PUT':
            return [IsAuthenticated(), ChangeReimbursementRequest()]
        return [IsAuthenticated()]

    def get(self, request):
        user = request.user
        queryset = Reimbursement.objects.all().order_by('-created_at')

        # Get filters
        area_manager_ids = request.query_params.getlist("area_manager")
        store_ids = request.query_params.getlist("stores")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        status = request.query_params.get("status")
        internal_control_status = request.query_params.get("internal_control_status", None)
        search = request.query_params.get("search")
        search_query = request.query_params.get("q", "").strip()
        region_id = request.query_params.get("region")
        disbursement_status = request.query_params.get("disbursement_status")

        # Role-based access
        if user.role.name == 'Restaurant Manager':
            queryset = queryset.filter(store_id=user.store_id)
        elif user.role.name == 'Area Manager':
            queryset = queryset.filter(store__in=user.assigned_stores.all())
        elif user.role.name == 'Internal Control':
            queryset = queryset.filter(
                Q(status__in=['approved']) &
                Q(Q(internal_control=user) | Q(internal_control__isnull=True)))
        elif user.role.name == 'Treasurer':
            queryset = queryset.filter(internal_control_status='approved')

        # Determine status field for this role
        if user.role.name == 'Treasurer':
            status_field = 'disbursement_status'
        elif user.role.name == 'Internal Control':
            status_field = 'internal_control_status'
        else:
            status_field = 'status'

        # Validate search_query before filters
        if search_query and not search_query.upper().startswith("RR-"):
            return CustomResponse(False, "Only RR-XXXX search is supported in 'q'", 400)

        def apply_common_filters(qs):
            if area_manager_ids:
                qs = qs.filter(store__area_manager__id__in=area_manager_ids)
            if store_ids:
                qs = qs.filter(store_id__in=store_ids)
            if region_id:
                qs = qs.filter(store__region_id=region_id)
            if start_date:
                try:
                    qs = qs.filter(created_at__date__gte=datetime.strptime(start_date, "%Y-%m-%d").date())
                except ValueError:
                    pass
            if end_date:
                try:
                    qs = qs.filter(created_at__date__lte=datetime.strptime(end_date, "%Y-%m-%d").date())
                except ValueError:
                    pass
            if disbursement_status:
                qs = qs.filter(disbursement_status=disbursement_status)
            if search:
                qs = qs.filter(
                    Q(requester__first_name__icontains=search) |
                    Q(requester__last_name__icontains=search)
                )
            if search_query:
                try:
                    request_id = int(search_query.upper().replace("RR-", ""))
                    qs = qs.filter(id=request_id)
                except ValueError:
                    pass
            return qs

        # count_queryset follows all filters EXCEPT status — feeds the tab counts
        count_queryset = apply_common_filters(queryset)

        # queryset follows ALL filters INCLUDING status — feeds the results list
        queryset = apply_common_filters(queryset)
        if status:
            queryset = queryset.filter(status=status)
        if internal_control_status:
            if internal_control_status in ["declined", "approved"]:
                queryset = queryset.filter(
                    internal_control_status=internal_control_status, internal_control=user)
            else:
                queryset = queryset.filter(internal_control_status=internal_control_status)

        # Status counts
        status_counts_all = (
            count_queryset
            .values(status_field)
            .annotate(count=Count(status_field))
            .order_by()
        )
        status_count_dict = {item[status_field]: item["count"] for item in status_counts_all}

        # Pagination and serialization
        paginator = DynamicPageSizePagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)
        serializer = ReimbursementSerializer(paginated_queryset, many=True)

        return CustomResponse(
            True,
            "Filtered reimbursement requests retrieved",
            200,
            {
                "count": paginator.page.paginator.count,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "results": serializer.data,
                "status_counts": status_count_dict,
            },
        )        

    def post(self, request):
        # Step 1: Create reimbursement (draft by default)
        print("Request Data ==> ", request.data)
        try:
            serializer = ReimbursementSerializer(data=request.data, context={'request': request})
            if not serializer.is_valid():
                errors = serializer.errors
                # Try to extract a meaningful message
                message = "Invalid data"
                if isinstance(errors, dict):
                    if 'detail' in errors:
                        message = errors['detail']

                    else:
                        first_key = next(iter(errors))
                        first_error = errors[first_key]
                        if isinstance(first_error, list):
                            message = first_error[0]
                        else:
                            message = first_error

                return CustomResponse(
                    False,
                    message,
                    400,
                    None
                )
            
            reimbursement = serializer.save(requester=request.user)

            # Update the related purchase request with the newly created reimbursement id
            purchase_request_refs = set()

            for item in reimbursement.items.all():
                ref = (item.purchase_request_ref or "").strip()
                match = re.match(r'^PR-0*(\d+)', ref)  # handles cases like PR-0015 or PR-0015-12000.00
                if match:
                    pr_id = int(match.group(1))  # convert to int (15)
                    print(pr_id)
                    purchase_request_refs.add(pr_id)

            if purchase_request_refs:
                purchase_requests = PurchaseRequest.objects.filter(id__in=purchase_request_refs)
                purchase_requests.update(reimbursement=reimbursement)
            
        # Sync receipt_validated from purchase request items to reimbursement items
            for item in reimbursement.items.all():
                pr_item_ref = (item.purchase_request_ref or "").strip()
                match = re.match(r'^PR-0*(\d+)', pr_item_ref)

                if not match:
                    continue

                pr_id = int(match.group(1))

                pr_item = (
                    PurchaseRequestItem.objects
                    .filter(
                        request_id=pr_id,
                    )
                    .order_by('-id')
                    .first()
                )

                if not pr_item:
                    continue

                item.receipt_validated = pr_item.receipt_validated
                item.save(update_fields=['receipt_validated'])

            return CustomResponse(
                True,
                "Reimbursement submitted for approval",
                201,
                ReimbursementSerializer(reimbursement).data
            )
        
        except Exception as err:
            return CustomResponse(
                False,
                "Unable to create reimbursement",
                400,
                {
                    "error":str(err)
                }
            )
        
        

    def put(self, request, pk, item_id=None):
        reimbursement = get_object_or_404(Reimbursement, pk=pk)
        serializer = ReimbursementUpdateSerializer(
            reimbursement, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            print(serializer.is_valid())
            serializer.save()
            return CustomResponse(True, "Reimbursement Request Updated Successfully", 200, serializer.data)
        return CustomResponse(False, serializer.errors, 400)


def extract_cloudinary_public_id(url: str) -> str:
    """
    Extract Cloudinary public_id (with folder, without extension) from a secure URL.
    
    Examples:
      https://res.cloudinary.com/<cloud>/image/upload/v123/receipts/myfile.jpg  → receipts/myfile
      https://res.cloudinary.com/<cloud>/raw/upload/v123/receipts/myfile.pdf   → receipts/myfile
    """
    match = re.search(r'/upload/(?:v\d+/)?(.+?)(\.[^./]+)?$', url)
    if match:
        return match.group(1)
    raise ValueError(f"Cannot extract public_id from Cloudinary URL: {url}")


def destroy_cloudinary_asset(public_id: str) -> bool:
    """
    Attempt to delete a Cloudinary asset trying both image and raw resource types.
    Returns True if deletion succeeded, False otherwise.
    PDFs are uploaded as resource_type='raw' by Cloudinary even when using 'auto',
    so we need to try both.
    """
    for resource_type in ("image", "raw"):
        try:
            result = cloudinary.uploader.destroy(public_id, resource_type=resource_type)
            if result.get("result") == "ok":
                return True
        except Exception:
            continue
    return False


class UploadReceiptView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, SubmitReimbursementRequest]

    def post(self, request):
        if 'receipt' not in request.FILES:
            return CustomResponse(False, "No receipt file provided.", 400)
        receipt_file = request.FILES['receipt']

        # --- Allow images AND PDFs ---
        ALLOWED_CONTENT_TYPES = [
            "image/jpeg", "image/png", "image/jpg", "image/webp",
            "application/pdf",
        ]
        if receipt_file.content_type not in ALLOWED_CONTENT_TYPES:
            return CustomResponse(
                False,
                "Invalid file type. Only JPEG, PNG, WebP images and PDF files are accepted.",
                400,
            )

        item_id = request.data.get('item_id')
        if not item_id:
            return CustomResponse(False, "Item ID is required for validation.", 400)

        try:
            item = PurchaseRequestItem.objects.get(id=item_id)
        except PurchaseRequestItem.DoesNotExist:
            return CustomResponse(False, "Invalid item ID.", 400)


        receipt_data = receipt_file.read()
        receipt_file.seek(0)
        is_pdf = receipt_file.content_type == "application/pdf"

        if getattr(settings, "ENVIRONMENT", "development") == "production":
            upload_kwargs = {
                "folder": "receipts/",
                "resource_type": "raw" if is_pdf else "image",
                "type": "upload",
                "access_mode": "public",
            }
            result = cloudinary.uploader.upload(receipt_file, **upload_kwargs)
            receipt_url = result.get("secure_url")
        else:
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'receipts/'))
            filename = fs.save(receipt_file.name, receipt_file)
            receipt_url = request.build_absolute_uri(settings.MEDIA_URL + 'receipts/' + filename)

        was_replacement = item.receipt_validated  # capture state BEFORE updating

        # Persist the upload — extraction is disabled for now but
        # we still need to mark the item as having a receipt.
        item.receipt_validated = True
        item.validation_errors = None
        item.save()

        return CustomResponse(
            True,
            "Receipt replaced successfully." if was_replacement else "Receipt uploaded successfully.",
            200,
            {"receipt_url": receipt_url}
        )
        
class ApproveReimbursementView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, ApproveReimbursementRequest]

    def post(self, request, pk):
       with transaction.atomic(): 
            reimbursement = get_object_or_404(Reimbursement, pk=pk)
            # self.check_object_permissions(request, reimbursement)
            role = request.user.role.name
            print("role", role)
            if role == "Area Manager":
                if reimbursement.status != "pending":
                    return CustomResponse(False, "Reimbursement is not pending", 400)

                reimbursement.status = "approved"
                reimbursement.area_manager = request.user
                reimbursement.area_manager_approved_at = timezone.now()
                reimbursement.items.update(status="approved")

            elif role == "Internal Control":
                if reimbursement.internal_control_status != "pending":
                    return CustomResponse(False, "Reimbursement is not pending", 400)

                reimbursement.internal_control_status = "approved"
                reimbursement.internal_control = request.user
                reimbursement.internal_control_approved_at = timezone.now()
                reimbursement.items.update(internal_control_status="approved")

            else:
                return CustomResponse(False, "Invalid role", 403)

            reimbursement.updated_by = request.user
            reimbursement.save(user=request.user)

            # Send notification to requester (AM) or Internal Control (IC)
            send_reimbursement_approval_notification(reimbursement, request.user)

            return CustomResponse(
                True,
                f"Reimbursement approved by {role} successfully",
                200
            )

class ApproveReimbursementItemView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, ApproveReimbursementRequest]

    def post(self, request, pk, item_id):
        with transaction.atomic():
            re = get_object_or_404(Reimbursement, pk=pk)
            item = get_object_or_404(ReimbursementItem, pk=item_id, reimbursement=re)
            items = re.items.all()

            self.check_object_permissions(request, re)

            role = request.user.role.name
            approved_now = False  # track if approval just completed

            if role == "Area Manager":
                if item.status == "approved":
                    return CustomResponse(False, "Item already approved", 400)

                item.status = "approved"
                item.save()

                if all(i.status == "approved" for i in items):
                    re.status = "approved"
                    re.area_manager = request.user
                    re.area_manager_approved_at = timezone.now()
                    approved_now = True

            elif role == "Internal Control":
                if item.internal_control_status == "approved":
                    return CustomResponse(False, "Item already approved", 400)

                item.internal_control_status = "approved"
                item.save()
                
                #Build item from reimbursements data to complete byd structure

                if all(i.internal_control_status == "approved" for i in items):
                    re.internal_control_status = "approved"
                    re.internal_control = request.user
                    re.internal_control_approved_at = timezone.now()
                    approved_now = True

            else:
                return CustomResponse(False, "Invalid role", 403)

            re.updated_by = request.user
            re.save(user=request.user)

            # SEND EMAIL ONLY WHEN FULL APPROVAL JUST HAPPENED
            if approved_now:
                send_reimbursement_approval_notification(
                    re,
                    request.user
                )

            return CustomResponse(
                True,
                {
                    "status": re.status,
                    "internal_control_status": re.internal_control_status,
                    "item_id": item.id
                },
                200
            )


class DeclineReimbursementView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, DeclineReimbursementRequest]

    def post(self, request, pk):
        comment_text = request.data.get("comment", "").strip()
        if not comment_text:
            return CustomResponse(False, "Comment is required when declining", 400)

        with transaction.atomic():
            re = (
                Reimbursement.objects
                .select_for_update()
                .get(pk=pk)
            )

            items = re.items.select_for_update().all()

            self.check_object_permissions(request, re)
            user_role = request.user.role.name

            # -------- INTERNAL CONTROL DECLINE --------
            if user_role == "Internal Control":
                re.internal_control_status = "declined"
                re.status = "pending"
                re.internal_control = request.user
                re.internal_control_declined_at = timezone.now()

                items.update(
                    status="pending",
                    internal_control_status="declined"
                )

                re.save(user=request.user)
                send_reimbursement_rejection_notification(
                    re, request.user, comment_text
                )
            # -------- AREA MANAGER FINAL DECLINE --------
            elif user_role == "Area Manager":
                re.status = "declined"
                re.area_manager = request.user
                re.area_manager_declined_at = timezone.now()
                items.update(status="declined")
                re.save(user=request.user)
                send_reimbursement_rejection_notification(
                    re, request.user, comment_text
                )

            else:
                return CustomResponse(False, "Invalid role", 403)

            ReimbursementComment.objects.create(
                reimbursement=re,
                author=request.user,
                text=comment_text
            )

        return CustomResponse(
            True,
            {
                "status": re.status,
                "internal_control_status": re.internal_control_status,
                "comment": comment_text
            },
            200
        )
    
class BulkApproveDeclineView(APIView):
    """Bulk approve or decline reimbursements. """

    queryset = Reimbursement.objects.all()
    serializer_class = ReimbursementSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthenticationFromCookie]

    def post(self, request):
        try:
            reimbursement_ids = request.data.get("reimbursement_ids", [])
            action = request.query_params.get("action", None)
            print("action ==> ", action)
            if not reimbursement_ids:
                return CustomResponse(
                    status=400,
                    valid=False,
                    msg=f"No Reimbursements selected for bulk update."
                )

            if not action:
                return CustomResponse(
                    status=400,
                    valid=False,
                    msg=f"Action is required. action must be either `approve` or `decline`"
                )

            if not action in ["approve", "decline"]:
                return CustomResponse(
                    status=400,
                    valid=False,
                    msg=f"Invalid action: {action}"
                )
            
            user = request.user
           
            reimbursements = self.queryset.filter(id__in=reimbursement_ids)
            reimbursements_array =[]
            reimbursement_items_array = []

            # check if any reimbursement exists for the specified IDs
            if reimbursements.exists():
                user_role = user.role.name
                print("user role ==> ", user_role)
                if user_role == Role.Type.AREA_MANAGER:
                    for reimbursement in reimbursements:
                        reimbursement.status = "approved" if action == "approve" else "declined"
                        reimbursement.area_manager=user
                        reimbursement.area_manager_approved_at=timezone.now() if action == "approve" else None
                        reimbursement.area_manager_declined_at=timezone.now() if action == "decline" else None
                        reimbursements_array.append(reimbursement)

                        for reimbursement_item in reimbursement.items.all():
                            reimbursement_item.status = "approved" if action == "approve" else "declined"
                            reimbursement_items_array.append(reimbursement_item)

                elif user_role == Role.Type.INTERNAL_CONTROL:
                    print("internal control ", user)
                    for reimbursement in reimbursements:
                        reimbursement.status="pending" if action == "decline" else reimbursement.status
                        reimbursement.internal_control=user
                        reimbursement.internal_control_status = "approved" if action == "approve" else "decline"
                        reimbursement.internal_control_approved_at=timezone.now() if action == "approve" else None
                        reimbursement.internal_control_declined_at=timezone.now() if action == "decline" else None
                        reimbursements_array.append(reimbursement)
                        print("reimbursement =>", reimbursement)
                        for reimbursement_item in reimbursement.items.all():
                            reimbursement.status="pending" if action == "decline" else reimbursement.status
                            reimbursement_item.internal_control_status= "approved" if action == "approve" else "decline"
                            reimbursement_items_array.append(reimbursement_item)
                            print("item ==> ", reimbursement_item)
                else:
                    return CustomResponse(
                        valid=False,
                        msg=f"You are not authorized to perform this action",
                        status=403
                    )
                # define internal control fields
                internal_control_fiels = [
                    "status",
                    "internal_control",
                    "internal_control_status",
                    "internal_control_approved_at",
                    "internal_control_declined_at"]
                # define area manager's fields
                area_manager_fields = [
                    "status",
                    "area_manager",
                    "area_manager_approved_at",
                    "area_manager_declined_at"]
                
                # Check fields to update
                fields_to_update = area_manager_fields if user_role == Role.Type.AREA_MANAGER else internal_control_fiels

                # Ensure the database is in a consistent state in event of 
                # any update failure.
                with transaction.atomic():
                    print("bulk updating... ", reimbursements_array)
                    Reimbursement.objects.bulk_update(reimbursements_array, 
                                                      fields=fields_to_update,
                                                      batch_size=200)
                    print("bulk updating items...", reimbursement_items_array)
                    ReimbursementItem.objects.bulk_update(reimbursement_items_array, 
                                                          fields=["status","internal_control_status"],
                                                          batch_size=200)
                    print("done bulk updating... ")
                    return CustomResponse(
                        valid=True,
                        msg=f"""{len(reimbursements_array)} reimbursements successfully {"approved" if action == "approve" else "declined"}""",
                        status=200
                    )

            else:
                return CustomResponse(
                    valid=False,
                    msg="Reimbursements do not exist",
                    status=400
                )
                
        except Exception as err:
            return CustomResponse(
                valid=False,
                msg=f"Unable to {action} reimbursement",
                data={
                    "error":str(err)
                }
            )


class DeclineReimbursementItemView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, DeclineReimbursementRequest]

    def post(self, request, pk, item_id):
        comment_text = request.data.get("comment", "").strip()
        if not comment_text:
            return CustomResponse(False, "Comment is required", 400)

        with transaction.atomic():
            re = (
                Reimbursement.objects
                .select_for_update()
                .get(pk=pk)
            )

            item = (
                ReimbursementItem.objects
                .select_for_update()
                .get(pk=item_id, reimbursement=re)
            )

            self.check_object_permissions(request, re)

            # ---- AREA MANAGER FLOW ----
            if request.user.role.name == "Area Manager":
                if item.status == "declined":
                    return CustomResponse(False, "Item already declined", 400)

                item.status = "declined"
                item.save()

                items = re.items.all()

                if items.filter(status="pending").exists():
                    re.status = "pending"
                else:
                    re.status = "declined"
                    re.area_manager = request.user
                    re.area_manager_declined_at = timezone.now()

                re.save(user=request.user)

                # Send email ONLY if final decline
                if re.status == "declined":
                    send_reimbursement_rejection_notification(
                        re, request.user, comment_text
                    )

            # ---- INTERNAL CONTROL FLOW ----
            elif request.user.role.name == "Internal Control":
                if item.internal_control_status == "declined":
                    return CustomResponse(False, "Item already declined", 400)

                item.internal_control_status = "declined"
                item.save()

                items = re.items.all()

                if items.filter(internal_control_status="pending").exists():
                    re.internal_control_status = "pending"
                else:
                    re.internal_control_status = "declined"
                    re.status = "pending"
                    re.internal_control = request.user
                    re.internal_control_declined_at = timezone.now()

                re.save(user=request.user)

                send_reimbursement_rejection_notification(
                    re, request.user, comment_text
                )

            ReimbursementComment.objects.create(
                reimbursement=re,
                author=request.user,
                text=comment_text
            )

        return CustomResponse(True, {
            "status": re.status,
            "item_id": item.id
        }, 200)
    
class ExportReimbursement(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, ViewReimbursementRequest]
    
    # -------------------------------
    # Helpers
    # -------------------------------
    #Helper methods called by the export view
    def get_queryset(self, user, start_date, end_date, status):
        qs = Reimbursement.objects.all()
        
        if user.role.name == "Area Manager":
            return qs.filter(
                store__in=user.assigned_stores.all(),
                created_at__date__range=(start_date, end_date),
                status__iexact=status,
            )

        if user.role.name == "Internal Control":
            return qs.filter(
                status="approved",
                created_at__date__range=(start_date, end_date),
                internal_control_status__iexact=status,
            )

        if user.role.name == "Treasurer":
            return qs.filter(
                internal_control_status="approved",
                created_at__date__range=(start_date, end_date),
                disbursement_status__iexact=status,
            )

        if user.role.name == "Restaurant Manager":
            return qs.filter(
                requester=user,
                created_at__date__range=(start_date, end_date),
                status__iexact=status,
            )

        return None

    def export_internal_control(self, queryset, start_date, end_date):
    
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Internal Control"

        internal_control_headers=[
            "Request ID",
            "Requester",
            "Store",
            "Store Code",
            "Store Manager",
            "Expense Item",
            "Amount"
            "Status"
            "Date Created"
        ]

        # expense_types = sorted({
        #     item.item_name
        #     for rr in queryset
        #     for item in rr.items.all()
        # })

        # headers = ["Staff Name"] + expense_types + ["Total"]
        sheet.append(internal_control_headers)
        book=[]

        for rr in queryset:
            store = rr.store
            store_name = store.name
            
            row = [
                f"RR-{rr.id:04d}",
                rr.requester.get_full_name(),
                store_name,
                store.code,
                store.area_manager.get_full_name() if store.area_manager else "Unknown",
                ",".join(rr.items.values_list("item_name", flat=True)),
                float(rr.total_amount),
                rr.internal_control_status,
                rr.created_at.strftime("%d-%m-%Y")
            ]
        
            sheet.append(row)
     
        return self.build_response(
            workbook,
            f"IC_reimbursements_{start_date.date()}_{end_date.date()}.xlsx"
        )

    def export_treasury(self, queryset:Reimbursement, start_date, end_date):
        from collections import defaultdict
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Treasury"

        treasurer_headers = [
            "Request ID",
            "Requester",
            "Region",
            "Store",
            "Store Code",
            "Expense Item",
            "Area Manager",
            "Amount",
            "Status",
            "Date Created",
            "Bank Account",
            "Bank GL Code"
        ]

        sheet.append(treasurer_headers)

        for rr in queryset:
            store = rr.store
            store_name = store.name

            row = [
                f"RR-{rr.id:04d}",
                rr.requester.get_full_name(),
                store.region.name if rr.store.region else "",
                store_name,
                store.code,
                ",".join(rr.items.values_list("item_name", flat=True)),
                store.area_manager.get_full_name() if store.area_manager else "Unknown",
                float(rr.total_amount),
                rr.status,
                rr.created_at.strftime("%d-%m-%Y"),
                rr.bank.bank_name if rr.bank else "Unknown",
                rr.bank.gl_code if rr.bank else "Unknown"
            ]
          
            sheet.append(row)

        return self.build_response(
            workbook,
            f"Treasury_reimbursements_{start_date.date()}_{end_date.date()}.xlsx"
        )


    def export_default(self, queryset, user, start_date, end_date):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        headers = ["Request ID", "Requester", "Store", "Total", "Status", "Date"]
        sheet.append(headers)

        for rr in queryset:
            sheet.append([
                f"RR-{rr.id:04d}",
                f"{rr.requester.first_name} {rr.requester.last_name}",
                rr.store.name if rr.store else "",
                rr.total_amount,
                rr.status.capitalize(),
                rr.created_at.strftime("%Y-%m-%d")
            ])

        return self.build_response(
            workbook,
            f"reimbursements_{start_date.date()}_{end_date.date()}.xlsx"
        )


    def build_response(self, workbook, filename):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    #-------------------------------
    # Main GET method for export
    #-------------------------------    
     
    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        status = request.query_params.get("status")
        # template = request.query_params.get("template") # e.g., "internal_control", "treasury", etc.

        user = request.user

        if not start_date or not end_date or not status:
            return CustomResponse(False, "start_date, end_date and status are required", 400)

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            return CustomResponse(False, "Invalid date format. Use YYYY-MM-DD", 400)

        if start_date > end_date:
            return CustomResponse(False, "start_date cannot be after end_date", 400)

        queryset = self.get_queryset(user, start_date, end_date, status)
        if queryset is None:
            return CustomResponse(False, "You are not allowed to export reimbursements", 403)

        if user.role.name == "Internal Control":
            return self.export_internal_control(queryset, start_date, end_date)

        if user.role.name == "Treasurer":
            return self.export_treasury(queryset, start_date, end_date)

        return self.export_default(queryset, user, start_date, end_date)




class DisbursemntView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, DisburseReimbursementRequest]
    
    # Disburse a reimbursement request and its items
    def post(self, request, pk):
        try:
            bank_id = request.data.get('bank', None)
            # account_id = request.data.get('account', None)

            if not bank_id:
                return CustomResponse(False, "Bank and account IDs are required", 400)

            reimbursement = get_object_or_404(Reimbursement, pk=pk)
            if reimbursement.disbursement_status != 'pending':
                return CustomResponse(False, "The selected reimbursement is not a pending disbursement", 400)
            
            reimbursement.bank = get_object_or_404(Bank, pk=bank_id)
            # reimbursement.account = get_object_or_404(Account, pk=account_id)
            
            reimbursement.disbursement_status = 'disbursed'
            reimbursement.treasurer = request.user
            reimbursement.disbursed_at = timezone.now()
            reimbursement.updated_by = request.user
            reimbursement.save(user=request.user)

            # Payload to be posted to SAP
            is_posted = update_sap_record(reimbursements=[reimbursement])
            print("is posted")
            if is_posted:
                logger.info("Reimbursement update successfully posted to BYD")
            else:
                logger.warning("Failed to post reimbursement update to BYD.")

            # UPDATE STORE BALANCE
            message = f"Reimbursement disbursed by Treasurer successfully"
            return CustomResponse(True, message, 200)
        
        except Exception as err:
            return CustomResponse(False, "Unable to disburse expense", 400, {"error":str(err)})
    
class BulkDisbursementView(APIView):
    authentication_classes = [JWTAuthenticationFromCookie]
    permission_classes = [IsAuthenticated, DisburseReimbursementRequest]
    
    # Bulk Disburse reimbursement requests and their items
    #ids refers to id of the selected reimbursement requests
    def post(self, request):
        ids = request.data.get('reimbursement_ids', [])
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return CustomResponse(False, "Invalid 'ids' format. Must be a list of integers.", 400)
        
        reimbursements = Reimbursement.objects.filter(id__in=ids)
        updated_count = 0
        reimbursements_data = []
        
        for reimbursement in reimbursements:
            if reimbursement.disbursement_status != 'pending':
                continue  # skip non-pending disbursements
            reimbursement.disbursement_status = 'disbursed'
            reimbursement.treasurer = request.user
            bank_id = request.data.get('bank')
            # account_id = request.data.get('account')
            reimbursement.bank = get_object_or_404(Bank, id=bank_id)
            # reimbursement.account = get_object_or_404(Account, id=account_id)
            reimbursement.disbursed_at = timezone.now()
            reimbursement.updated_by = request.user
            reimbursement.save(user=request.user)
            reimbursements_data.append(reimbursement)
            updated_count += 1
            
        is_posted = update_sap_record(reimbursements=reimbursements_data)
        print("is posted", is_posted, reimbursements_data)
        if is_posted:
                logger.info("Reimbursement update successfully posted to BYD")
        else:
            logger.warning("Failed to post reimbursement update to BYD.")
        
        return CustomResponse(True, f"{updated_count} reimbursement(s) disbursed successfully", 200)
    
    
